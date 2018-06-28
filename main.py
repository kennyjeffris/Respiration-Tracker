import serial
import serial.tools.list_ports
from tkinter.filedialog import asksaveasfilename
import time
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font 

# Setup Sequence
def run_setup(gui):
    # Assign workbook to gui for future use.
    gui.workbook = prepare_workbook()
    # Check to see if the interval is valid, else will not proceed.
    interval = gui.return_interval()
    # Return false if something doesn't work so that loop does not begin.
    proceed = False
    if interval:
        # Assign arduino object to gui for future use.
        gui.arduino = attach_arduino(gui)
        if gui.arduino:
            # Choose file save path
            gui.filename = prepare_save(gui)
            if gui.filename:
                # If everything is good, proceed.
                proceed = True
                # Set button to be interrupt.
                gui.update_button('Finish', True)
        else:
            gui.update_button('Retry', False)
    else:
        gui.update_button('Retry', False)
    return proceed

def run_loop(gui):
    # Start time of acquisition for timing purposes.
    gui.start_time = time.time()
    # Interval time is in minutes - convert to seconds.
    interval = gui.return_interval() * 60
    # Begin adding new lines to workbook at line 2 (1 is descriptors).
    index = 2
    # Get first reading right away
    firstTime = True
    # Only continue if Finish button has not been pressed.
    while not gui.stop_thread.is_set():
        if (time.time() - gui.start_time > (interval - 30) or firstTime):
            # Print statement for debugging later (commented out now).
            #print('Starting Acquisition at time {}'.format(time.time() - gui.start_time))
            # Disable Finish button until it after acquisition to avoid crash:
            gui.disable_button()
            # Get the calculated respiration rate.
            rate = acquire_data(gui)
            # Aded timestamp and rate to workbook.
            gui.workbook = update_workbook(gui, rate, index)
            # Save everytime it is added in case of failure.
            save_file(gui.workbook, gui.filename)
            # Update line to add entries to.
            index += 1
            if firstTime:
                # Reset start time to make sure the next reading happens at the correct interval.
                gui.start_time = time.time()
                firstTime = False
            else: 
                # Increment interval so that the time-check still works.
                interval += (gui.return_interval() * 60)
            gui.update_label('Waiting until next acquisition...')
            gui.enable_button()

def update_workbook(gui, rate, index):
    # Add lines to workbook
    wb = gui.workbook
    ws = wb.worksheets[0]
    a = ws.cell(row=index, column=1, value=datetime.datetime.now().strftime("%H:%M:%S"))
    b = ws.cell(row=index, column=2, value=rate)
    return wb




def attach_arduino(gui):
    # Choose correct port (Code from https://stackoverflow.com/questions/24214643/python-to-automatically-select-serial-ports-for-arduino)
    arduino_ports = [
        p.device
        for p in serial.tools.list_ports.comports()
        if 'Arduino' in p.description
    ]
    if not arduino_ports:
        gui.update_label('No Arduinos found, please connect one and retry.')
    elif len(arduino_ports) > 1:
        gui.update_label('Multiple Arduinos found - using the first.')
    else:
        gui.update_label('Arduino found at port {}'.format(arduino_ports[0]))
        ser = serial.Serial(
            port = arduino_ports[0],
            baudrate=115200,
            )
        # Wait 1.5 Seconds before communicating because connection causes arduino to reset.
        time.sleep(3)
        ser.write('2'.encode())
        return ser

def prepare_workbook():
    # Initialize excel workbook for saving data
    wb = Workbook()
    main_page = wb.worksheets[0]
    a1 = main_page.cell(row=1, column=1, value="TimeStamp")
    b1 = main_page.cell(row=1, column=2, value="Respiration Rate (breaths/min)")

    a1.font = b1.font = Font(bold=True)
    return wb

def prepare_save(gui):
    # Prepare filename
    delim = "_"
    date_time = datetime.datetime.now().strftime("%Y_%m_%d-%H_%M")
    file_parts = [date_time, 'respiration', 'data']
    newName = delim.join(file_parts)
    options = {}
    options['defaultextension'] = ".xlsx"
    options['filetypes'] = (('xlsx files', '*.xlsx'), ('all files', '*.*'))
    options['initialfile'] = newName
    options['title'] = "Save as..."
    dest_filename = asksaveasfilename(**options)
    if not dest_filename:
        # Let users know the cancelled save.
        gui.update_label('No save location selected, please retry.')
        gui.update_button('Retry', False)
    else:
        return dest_filename

def save_file(wb, dest_filename):
    try:
        # Save file
        wb.save(filename=dest_filename)
    except PermissionError as e:
        pass
    if not dest_filename:
        exit()

def acquire_data(gui):
    # Code adapted from https://gist.github.com/brandoncurtis/33a67d9d402973face8d
    gui.update_label('Obtaining Respiration Data, Do Not Disturb Tubing.')
    ser = gui.arduino
    run = True
    # Get start time of acquisition
    start_time = time.time()
    # Initialize arrays
    data = []
    timepoints = []
    # 30 Second Intervals (breaths/min = num breaths in 30 sec * 2)
    duration = 30
    # Clean out junk from serial connection
    ser.flushInput()
    while run:
        # More serial connection maintenance
        ser.reset_input_buffer()
        # Get number and extract the number from it
        new_data = ser.readline().decode('utf-8')
        try:
            # Add to array
            data.append(float(new_data[0]))
            timepoints.append(time.time() - start_time)
            current_time = timepoints[-1]
            # Quit after duration is passed
            if timepoints[-1] > duration: run=False
        except: pass
    filtered_time = []
    # Get the number of breaths
    for index, value in enumerate(data):
        # Prevent OutofBounds error
        if index == 0:
            pass
            # Only count nonzero value that is preceeded by 0 (start of breath)
        if value > 0 and data[index - 1] == 0:
            filtered_time.append(timepoints[index])
    # Return value times 2 because we need total breaths per minute
    return len(filtered_time) * 2
    